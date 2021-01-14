import openpyxl
wb = openpyxl.load_workbook('SAFPRO Supplier 2021.xlsx')
sheet1 = wb['Sheet1']

wb2 = openpyxl.load_workbook('SAFPRORawFish2021.xlsx')
sheet = wb2['Sheet2']
count = 1
for col_cells in sheet1.iter_cols(min_col=1, max_col=1):
    for cell in col_cells:
        coor1 = 'G' + str(cell.row)
        coor2 = 'A' + str(cell.row - 1)
        if(cell.value == '2300/010' and (sheet1[coor1]).value != "" and 'transport' not in ((sheet1[coor2]).value).lower()):

            coor = 'G' + str(cell.row) #getting co-or for amount 1
            amount1 = (sheet1[coor]).value
            coor = 'H' + str(cell.row)  # getting co-or for amount 2
            amount2 = (sheet1[coor]).value
            coor = 'I' + str(cell.row)  # getting co-or for amount 3
            amount3 = (sheet1[coor]).value

            coor = 'A' + str(cell.row - 1)
            supplier = (sheet1[coor]).value
            add1 = -2
            while("Supplier" not in str(supplier) and add1 != -10):
                coor = 'A' + str(cell.row - 1 + add1)
                supplier = (sheet1[coor]).value
                add1 = add1 - 2

            coor = 'A' + str(cell.row + 1)
            quan = (sheet1[coor]).value

            coor = 'A' + str(cell.row + 2)
            totalsupp = (sheet1[coor]).value
            coor = 'B' + str(cell.row+2)
            date = (sheet1[coor]).value

            coor = 'C' + str(cell.row+2)
            orderinv = (sheet1[coor]).value
            add2 = 2
            while("Total" not in str(totalsupp) and add2 != 10):

                coor = 'A' + str(cell.row + 2 + add2)
                totalsupp = (sheet1[coor]).value
                coor = 'B' + str(cell.row + 2 + add2)
                date = (sheet1[coor]).value
                coor = 'C' + str(cell.row + 2 + add2)
                orderinv = (sheet1[coor]).value
                add2 = add2 + 2

            coor = 'A' + str(count)
            sheet[coor] = date
            coor = 'B' + str(count)
            sheet[coor] = totalsupp
            coor = 'C' + str(count)
            sheet[coor] = orderinv
            coor = 'D' + str(count)
            sheet[coor] = 'RAW FISH'
            coor = 'E' + str(count)
            sheet[coor] = supplier
            coor = 'F' + str(count)
            sheet[coor] = quan
            coor = 'G' + str(count)
            sheet[coor] = amount1
            coor = 'H' + str(count)
            sheet[coor] = amount2
            coor = 'I' + str(count)
            sheet[coor] = amount3
            count = count + 1
            wb2.save('SAFPRORawFish2021.xlsx')